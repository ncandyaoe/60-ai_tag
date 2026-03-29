#!/usr/bin/env python3
"""
nut_config_excel.py — 营养表配置 Excel 双向同步工具
====================================================

Sheet 1: 布局总表     — NutritionLayout 全局参数
Sheet 2: 标题行配置   — NutHeaderRow 列表
Sheet 3: 列定义       — NutColumn 列表
Sheet 4: 数据行模板   — table_data 的显示属性模板

功能:
    export  — 将代码配置 + app.py 测试数据导出为 Excel
    load    — 从 Excel 读取，构建 NutritionLayout + 数据行模板

Usage:
    python nut_config_excel.py export           # 导出到 nut_config.xlsx
    python nut_config_excel.py load             # 验证加载
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from nut_layouts import NUT_LAYOUT_REGISTRY, NutritionLayout, NutHeaderRow, NutColumn
from typing import List, Dict, Optional

# ══════════════════════════════════════════════════════════
# 样式常量
# ══════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DESC_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DESC_FONT = Font(name="Arial", size=9, italic=True, color="333333")
DATA_FONT = Font(name="Arial", size=10)
COUNTRY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUB_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 子项浅黄底
HEAVY_FONT = Font(name="Arial", size=10, bold=True, color="8B0000")  # Heavy 行深红粗体
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ══════════════════════════════════════════════════════════
# Sheet 定义
# ══════════════════════════════════════════════════════════

LAYOUT_COLUMNS = [
    ("country_code",          "国家代码",               "如 CA, AU, US 等"),
    ("name",                  "布局名称",               "可读的描述名"),
    ("draw_data_row_lines",   "数据行横线",             "True/False：数据行之间是否画分隔线"),
    ("data_row_line_width",   "数据行线宽(pt)",         "默认0.3"),
    ("data_row_line_padding", "数据行线左右退让(pt)",    "横线两端向内缩进的距离"),
    ("header_line_width",     "标题区线宽(pt)",         "标题区域底部横线粗细"),
    ("border_line_width",     "边框线宽(pt)",           "整个营养表外框粗细"),
    ("outer_border_line_width","外框独立线宽(pt)",      "留空则等于 border_line_width"),
    ("thick_line_width",      "粗分割线宽(pt)",         "加粗分组线（如CA的粗横线）"),
    ("sub_indent",            "子项缩进(pt)",           "子项（如 Saturated）向右缩进距离"),
    ("draw_col_sep",          "列竖线",                 "True/False：是否画列之间的竖线"),
    ("col_sep_in_data",       "竖线延伸到数据区",       "False 则竖线只在标题区"),
    ("col_padding",           "列内边距(pt)",           "文本与竖线的左右留白"),
    ("line_height_ratio",     "行高比",                 "行高 = 字号 × 此值。增大=行距更宽"),
    ("bold_main_items",       "主项加粗",               "True = 非子项的数据行名称全部加粗"),
    ("reference_font_size",  "参考字号(pt)",           ">0时启用自适应缩放。所有pt绝对值按 实际字号/参考字号 缩放"),
]

HEADER_ROW_COLUMNS = [
    ("country_code",       "国家代码",               "对应布局总表的国家"),
    ("row_index",          "行序号",                 "从0开始的排列顺序"),
    ("cells",              "文字内容",               "单元格文本（多列用 | 分隔，换行用 \\n）"),
    ("bold",               "加粗",                   "True/False"),
    ("span_full",          "跨全行",                 "True = 合并所有列"),
    ("template",           "模板占位",               "True = 文本含 {key} 占位符，运行时替换"),
    ("multi_line",         "多行渲染",               "True = 单元格内允许\\n换行"),
    ("align",              "水平对齐",               "left / center / right"),
    ("valign",             "垂直对齐",               "center / ca_header（底部对齐）"),
    ("font_ratio",         "字号缩放比",             "1.0=基准。1.47=放大47%"),
    ("font_ratios",        "分列字号缩放",           "按列不同字号，如 1.5|0.77"),
    ("height_ratio",       "行高缩放比",             "1.0=标准。0.74=压扁到74%"),
    ("margin_top",         "顶部偏移(pt)",           "负值=往上吸附紧贴，正值=往下推开"),
    ("margin_below",       "底部偏移(pt)",           "行边框下方与下一行的额外留白"),
    ("draw_line_below",    "底部横线",               "True/False：本行下方是否画线"),
    ("line_width_below",   "底线倍率",               "0=标准线宽。负数=外框线宽的倍率（如-2.5=2.5倍粗线）"),
    ("line_span_col",      "底线跨列",               "留空=全宽。填列索引(0开始)=只画到该列右边"),
    ("line_left_padding",  "底线左退让(pt)",          "底部横线左端向右缩"),
    ("independent_tz",     "独立压缩",               "True = 单独计算横向压缩比"),
    ("horizontal_padding", "左右内距(pt)",            "文字与行框的左右预留距离"),
    ("font_override",      "覆盖字体",               "指定字体名，如 AliPuHuiTi-Heavy。多列用 | 分隔"),
    ("col_width_ratios",   "列宽覆盖",               "局部列宽比例，如 0.4|0.6"),
]

COLUMN_COLUMNS = [
    ("country_code", "国家代码",   "对应布局总表的国家"),
    ("col_index",    "列序号",     "从0开始"),
    ("key",          "数据键名",   "name / per_serving / per_100g / nrv / value"),
    ("width_ratio",  "列宽比例",   "所有列之和 = 1.0"),
    ("align",        "对齐方式",   "left / center / right"),
]

# Sheet 4: 数据行模板
DATA_ROW_COLUMNS = [
    ("country_code",       "国家代码",      "对应布局总表的国家"),
    ("row_index",          "行序号",        "从0开始，决定显示顺序"),
    ("name",               "营养素名称",    "双语名称如 Fat / Lipides。脚注用 * 开头"),
    ("heavy",              "Heavy加粗",     "True = 名称用最粗字体（Heavy），数值不加粗"),
    ("bold",               "Bold加粗",      "True/False。留空=跟随全局 bold_main_items"),
    ("is_sub",             "子项",          "True = 缩进渲染（如 Saturated 是 Fat 的子项）"),
    ("height_ratio",       "行高缩放",      "1.0=标准。0.8=压扁。0.72=脚注"),
    ("hide_line_below",    "隐藏底线",      "True = 不画此行下方的分隔线"),
    ("thick_line_below",   "粗底线",        "True = 此行下方画加粗分隔线"),
    ("padded_line_below",  "退让底线",      "True = 此行底线左右退让（分组呼吸感）"),
    ("margin_top",         "顶部偏移(pt)",  "正值=往下推开与上一行距离"),
]

# ══════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════

def _write_header(ws, columns_def):
    for ci, (field, cn_name, desc) in enumerate(columns_def, 1):
        cell1 = ws.cell(row=1, column=ci, value=cn_name)
        cell1.font = HEADER_FONT
        cell1.fill = HEADER_FILL
        cell1.alignment = Alignment(horizontal='center', wrap_text=True)
        cell1.border = THIN_BORDER
        
        cell2 = ws.cell(row=2, column=ci, value=desc)
        cell2.font = DESC_FONT
        cell2.fill = DESC_FILL
        cell2.alignment = Alignment(wrap_text=True)
        cell2.border = THIN_BORDER
        
        cell3 = ws.cell(row=3, column=ci, value=field)
        cell3.font = Font(name="Consolas", size=9, color="888888")
        cell3.border = THIN_BORDER
    
    ws.freeze_panes = "A4"
    return 4


def _serialize_list(val):
    if val is None:
        return ""
    if isinstance(val, list):
        return "|".join(str(v) for v in val)
    return str(val)


def _parse_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if val is None or val == "":
        return False
    if isinstance(val, str):
        return val.strip().lower() in ("true", "1", "yes", "是")
    return bool(val)


def _parse_bool_or_none(val):
    """解析布尔值，空值返回 None（用于 bold 字段的"留空=继承全局"语义）"""
    if val is None or val == "":
        return None
    return _parse_bool(val)


def _parse_float(val, default=0.0) -> float:
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _parse_optional_float(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_list_float(val):
    if val is None or val == "":
        return None
    parts = str(val).split("|")
    try:
        return [float(p.strip()) for p in parts if p.strip()]
    except ValueError:
        return None


def _parse_list_str(val):
    if val is None or val == "":
        return None
    parts = str(val).split("|")
    result = [p.strip() for p in parts if p.strip()]
    return result if len(result) > 1 else (result[0] if result else None)


def _parse_optional_int(val):
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _read_sheet_as_dicts(ws):
    headers = []
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=ci).value
        if v:
            headers.append(str(v).strip())
        else:
            headers.append(f"col_{ci}")
    
    rows = []
    for ri in range(4, ws.max_row + 1):
        row_data = {}
        all_empty = True
        for ci, key in enumerate(headers):
            val = ws.cell(row=ri, column=ci + 1).value
            row_data[key] = val
            if val is not None and val != "":
                all_empty = False
        if not all_empty:
            rows.append(row_data)
    return rows


# ══════════════════════════════════════════════════════════
# 内置测试数据（从 app.py 收集）
# ══════════════════════════════════════════════════════════

def _get_all_test_data():
    """收集 app.py 中各国测试数据的 table_data"""
    # 动态导入，避免循环依赖
    try:
        import importlib
        # 先获取已有测试数据
        test_data = {}
        
        # CA 数据
        test_data["CA"] = [
            {"name": "Fat / Lipides", "heavy": True, "hide_line_below": True, "height_ratio": 0.8},
            {"name": "Saturated / saturées", "is_sub": True, "hide_line_below": True, "height_ratio": 0.8},
            {"name": "+ Trans / trans", "is_sub": True, "padded_line_below": True, "height_ratio": 0.8},
            {"name": "Carbohydrate / Glucides", "hide_line_below": True, "height_ratio": 0.8},
            {"name": "Fibre / Fibres", "is_sub": True, "hide_line_below": True, "height_ratio": 0.8},
            {"name": "Sugars / Sucres", "is_sub": True, "padded_line_below": True, "height_ratio": 0.8},
            {"name": "Protein / Protéines", "heavy": True, "padded_line_below": True},
            {"name": "Cholesterol / Cholestérol", "heavy": True, "padded_line_below": True},
            {"name": "Sodium", "heavy": True, "thick_line_below": True},
            {"name": "Potassium"},
            {"name": "Calcium"},
            {"name": "Iron / Fer", "thick_line_below": True},
            {"name": "* 5% or less is a little, 15% or more is a lot", "bold": False, "hide_line_below": True, "height_ratio": 0.72, "margin_top": 2.0},
            {"name": "* 5% ou moins c'est peu, 15% ou plus c'est beaucoup", "bold": False, "hide_line_below": True, "height_ratio": 0.72},
        ]
        
        # AU 数据
        test_data["AU"] = [
            {"name": "Energy"},
            {"name": "Protein"},
            {"name": "Carbohydrate"},
            {"name": "of which total sugars", "is_sub": True},
            {"name": "Total fat"},
            {"name": "of which saturated fat", "is_sub": True},
            {"name": "Sodium"},
        ]
        
        # EU 通用数据
        test_data["EU_MULTI"] = [
            {"name": "Energy / Energie / Valor energético / Énergie"},
            {"name": "Fat / Vetten / Grasas / Fett / Matières grasses"},
            {"name": "of which / waarvan / de las cuales / davon / dont", "is_sub": True},
            {"name": "-Saturates / Verzadigde vetzuren / Saturadas / gesättigte Fettsäuren / Acides gras saturés", "is_sub": True},
            {"name": "Carbohydrate / Koolhydraten / Hidratos de carbono / Kohlenhydrate / Glucides"},
            {"name": "of which / waarvan / de las cuales / davon / dont", "is_sub": True},
            {"name": "-Sugars / Suikers / Azúcares / Zucker / Sucres", "is_sub": True},
            {"name": "Protein / Eiwitten / Proteínas / Eiweiß / Protéines"},
            {"name": "Salt / Zout / Sal / Salz / Sel"},
        ]
        
        return test_data
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════
# 导出
# ══════════════════════════════════════════════════════════

def export_to_excel(output_path: str = "nut_config.xlsx"):
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: 布局总表 ──
    ws_layout = wb.active
    ws_layout.title = "布局总表"
    start_row = _write_header(ws_layout, LAYOUT_COLUMNS)
    
    for row_idx, (code, layout) in enumerate(NUT_LAYOUT_REGISTRY.items()):
        r = start_row + row_idx
        values = [
            code, layout.name, layout.draw_data_row_lines,
            layout.data_row_line_width, layout.data_row_line_padding,
            layout.header_line_width, layout.border_line_width,
            layout.outer_border_line_width or "",
            layout.thick_line_width, layout.sub_indent, layout.draw_col_sep,
            layout.col_sep_in_data, layout.col_padding,
            layout.line_height_ratio, layout.bold_main_items,
            layout.reference_font_size or "",
        ]
        for ci, val in enumerate(values, 1):
            cell = ws_layout.cell(row=r, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if ci == 1:
                cell.fill = COUNTRY_FILL
                cell.font = Font(name="Arial", size=10, bold=True)
    
    for ci in range(1, len(LAYOUT_COLUMNS) + 1):
        ws_layout.column_dimensions[get_column_letter(ci)].width = 18
    
    # ── Sheet 2: 标题行配置 ──
    ws_header = wb.create_sheet("标题行配置")
    start_row = _write_header(ws_header, HEADER_ROW_COLUMNS)
    
    data_row = start_row
    for code, layout in NUT_LAYOUT_REGISTRY.items():
        for hi, hdr in enumerate(layout.header_rows):
            values = [
                code, hi,
                "|".join(hdr.cells).replace("\n", "\\n"),
                hdr.bold, hdr.span_full, hdr.template, hdr.multi_line,
                hdr.align, hdr.valign,
                hdr.font_ratio, _serialize_list(hdr.font_ratios),
                hdr.height_ratio, hdr.margin_top, hdr.margin_below,
                hdr.draw_line_below, hdr.line_width_below,
                hdr.line_span_col if hdr.line_span_col is not None else "",
                hdr.line_left_padding, hdr.independent_tz,
                hdr.horizontal_padding if hdr.horizontal_padding is not None else "",
                _serialize_list(hdr.font_override),
                _serialize_list(hdr.col_width_ratios),
            ]
            for ci, val in enumerate(values, 1):
                cell = ws_header.cell(row=data_row, column=ci, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if ci == 1:
                    cell.fill = COUNTRY_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
            data_row += 1
    
    for ci in range(1, len(HEADER_ROW_COLUMNS) + 1):
        ws_header.column_dimensions[get_column_letter(ci)].width = 16
    ws_header.column_dimensions["C"].width = 40
    
    # ── Sheet 3: 列定义 ──
    ws_col = wb.create_sheet("列定义")
    start_row = _write_header(ws_col, COLUMN_COLUMNS)
    
    data_row = start_row
    for code, layout in NUT_LAYOUT_REGISTRY.items():
        for ci_idx, col in enumerate(layout.columns):
            values = [code, ci_idx, col.key, col.width_ratio, col.align]
            for ci, val in enumerate(values, 1):
                cell = ws_col.cell(row=data_row, column=ci, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if ci == 1:
                    cell.fill = COUNTRY_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
            data_row += 1
    
    for ci in range(1, len(COLUMN_COLUMNS) + 1):
        ws_col.column_dimensions[get_column_letter(ci)].width = 16
    
    # ── Sheet 4: 数据行模板 ──
    ws_data = wb.create_sheet("数据行模板")
    start_row = _write_header(ws_data, DATA_ROW_COLUMNS)
    
    test_data = _get_all_test_data()
    data_row = start_row
    for code, rows in test_data.items():
        for ri, item in enumerate(rows):
            values = [
                code,
                ri,
                item.get("name", ""),
                item.get("heavy", False),
                item.get("bold", ""),   # 留空 = 继承全局
                item.get("is_sub", False),
                item.get("height_ratio", ""),  # 留空 = 默认 1.0
                item.get("hide_line_below", False),
                item.get("thick_line_below", False),
                item.get("padded_line_below", False),
                item.get("margin_top", ""),  # 留空 = 默认 0
            ]
            for ci, val in enumerate(values, 1):
                cell = ws_data.cell(row=data_row, column=ci, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                
                # 国家列绿底
                if ci == 1:
                    cell.fill = COUNTRY_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
                # 子项行浅黄底
                elif item.get("is_sub"):
                    cell.fill = SUB_FILL
                # Heavy 行名称列深红粗体
                if ci == 3 and item.get("heavy"):
                    cell.font = HEAVY_FONT
            
            data_row += 1
    
    # 列宽
    ws_data.column_dimensions["A"].width = 12
    ws_data.column_dimensions["B"].width = 8
    ws_data.column_dimensions["C"].width = 55  # 营养素名称宽一些
    for ci in range(4, len(DATA_ROW_COLUMNS) + 1):
        ws_data.column_dimensions[get_column_letter(ci)].width = 14
    
    wb.save(output_path)
    n_data_rows = sum(len(v) for v in test_data.values())
    print(f"✅ 已导出到 {output_path}")
    print(f"   - 布局总表: {len(NUT_LAYOUT_REGISTRY)} 个国家")
    print(f"   - 标题行配置: {sum(len(l.header_rows) for l in NUT_LAYOUT_REGISTRY.values())} 行")
    print(f"   - 列定义: {sum(len(l.columns) for l in NUT_LAYOUT_REGISTRY.values())} 列")
    print(f"   - 数据行模板: {n_data_rows} 行 ({len(test_data)} 个国家)")


# ══════════════════════════════════════════════════════════
# 导入
# ══════════════════════════════════════════════════════════

def load_from_excel(excel_path: str = "nut_config.xlsx"):
    """从 Excel 加载配置，返回 (registry, data_templates)"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # ── 列定义 ──
    col_data = _read_sheet_as_dicts(wb["列定义"])
    columns_by_country = {}
    for row in col_data:
        code = str(row["country_code"]).strip()
        if code not in columns_by_country:
            columns_by_country[code] = []
        columns_by_country[code].append(NutColumn(
            key=str(row["key"]),
            width_ratio=_parse_float(row["width_ratio"], 0.5),
            align=str(row.get("align", "center") or "center"),
        ))
    
    # ── 标题行 ──
    hdr_data = _read_sheet_as_dicts(wb["标题行配置"])
    headers_by_country = {}
    for row in hdr_data:
        code = str(row["country_code"]).strip()
        if code not in headers_by_country:
            headers_by_country[code] = []
        
        cells_raw = str(row.get("cells", "")).replace("\\n", "\n")
        cells = cells_raw.split("|") if "|" in cells_raw else [cells_raw]
        
        hdr = NutHeaderRow(
            cells=cells,
            bold=_parse_bool(row.get("bold", False)),
            span_full=_parse_bool(row.get("span_full", False)),
            template=_parse_bool(row.get("template", False)),
            multi_line=_parse_bool(row.get("multi_line", False)),
            align=str(row.get("align", "center") or "center"),
            valign=str(row.get("valign", "center") or "center"),
            font_ratio=_parse_float(row.get("font_ratio"), 1.0),
            font_ratios=_parse_list_float(row.get("font_ratios")),
            height_ratio=_parse_float(row.get("height_ratio"), 1.0),
            margin_top=_parse_float(row.get("margin_top"), 0.0),
            margin_below=_parse_float(row.get("margin_below"), 0.0),
            draw_line_below=_parse_bool(row.get("draw_line_below", True)),
            line_width_below=_parse_float(row.get("line_width_below"), 0.0),
            line_span_col=_parse_optional_int(row.get("line_span_col")),
            line_left_padding=_parse_float(row.get("line_left_padding"), 0.0),
            independent_tz=_parse_bool(row.get("independent_tz", False)),
            horizontal_padding=_parse_optional_float(row.get("horizontal_padding")),
            font_override=_parse_list_str(row.get("font_override")),
            col_width_ratios=_parse_list_float(row.get("col_width_ratios")),
        )
        headers_by_country[code].append(hdr)
    
    # ── 布局总表 ──
    layout_data = _read_sheet_as_dicts(wb["布局总表"])
    registry = {}
    for row in layout_data:
        code = str(row["country_code"]).strip()
        layout = NutritionLayout(
            name=str(row.get("name", code)),
            columns=columns_by_country.get(code, []),
            header_rows=headers_by_country.get(code, []),
            draw_data_row_lines=_parse_bool(row.get("draw_data_row_lines", True)),
            data_row_line_width=_parse_float(row.get("data_row_line_width"), 0.3),
            data_row_line_padding=_parse_float(row.get("data_row_line_padding"), 0.0),
            header_line_width=_parse_float(row.get("header_line_width"), 0.5),
            border_line_width=_parse_float(row.get("border_line_width"), 0.5),
            outer_border_line_width=_parse_optional_float(row.get("outer_border_line_width")),
            thick_line_width=_parse_float(row.get("thick_line_width"), 1.5),
            sub_indent=_parse_float(row.get("sub_indent"), 10.0),
            draw_col_sep=_parse_bool(row.get("draw_col_sep", True)),
            col_sep_in_data=_parse_bool(row.get("col_sep_in_data", True)),
            col_padding=_parse_float(row.get("col_padding"), 1.5),
            line_height_ratio=_parse_float(row.get("line_height_ratio"), 1.15),
            bold_main_items=_parse_bool(row.get("bold_main_items", False)),
            reference_font_size=_parse_float(row.get("reference_font_size"), 0.0),
        )
        registry[code] = layout
    
    # ── 数据行模板 ──
    data_templates = {}
    if "数据行模板" in wb.sheetnames:
        dt_data = _read_sheet_as_dicts(wb["数据行模板"])
        for row in dt_data:
            code = str(row["country_code"]).strip()
            if code not in data_templates:
                data_templates[code] = []
            
            item = {"name": str(row.get("name", ""))}
            
            if _parse_bool(row.get("heavy")):
                item["heavy"] = True
            
            bold_val = _parse_bool_or_none(row.get("bold"))
            if bold_val is not None:
                item["bold"] = bold_val
            
            if _parse_bool(row.get("is_sub")):
                item["is_sub"] = True
            
            hr = _parse_optional_float(row.get("height_ratio"))
            if hr is not None:
                item["height_ratio"] = hr
            
            if _parse_bool(row.get("hide_line_below")):
                item["hide_line_below"] = True
            
            if _parse_bool(row.get("thick_line_below")):
                item["thick_line_below"] = True
            
            if _parse_bool(row.get("padded_line_below")):
                item["padded_line_below"] = True
            
            mt = _parse_optional_float(row.get("margin_top"))
            if mt is not None and mt != 0:
                item["margin_top"] = mt
            
            data_templates[code].append(item)
    
    # ── 底部行配置 ──
    footer_by_country = {}
    if "底部行配置" in wb.sheetnames:
        from nut_layouts import NutFooterRow
        ftr_data = _read_sheet_as_dicts(wb["底部行配置"])
        for row in ftr_data:
            code = str(row.get("country_code", "")).strip()
            if not code:
                continue
            if code not in footer_by_country:
                footer_by_country[code] = []
            
            ftr = NutFooterRow(
                text=str(row.get("text", "")),
                bold=_parse_bool(row.get("bold", False)),
                font_ratio=_parse_float(row.get("font_ratio"), 0.8),
                height_ratio=_parse_float(row.get("height_ratio"), 1.6),
                draw_line_below=_parse_bool(row.get("draw_line_below", True)),
                thick_line_below=_parse_bool(row.get("thick_line_below", False)),
                margin_top=_parse_float(row.get("margin_top"), 0.0),
                align=str(row.get("align", "left") or "left"),
            )
            footer_by_country[code].append(ftr)
    
    # 挂载 footer_rows 到各国布局
    for code, footers in footer_by_country.items():
        if code in registry:
            registry[code].footer_rows = footers

    wb.close()
    n_dt = sum(len(v) for v in data_templates.values())
    n_ft = sum(len(v) for v in footer_by_country.values())
    print(f"✅ 从 {excel_path} 加载了 {len(registry)} 个国家的配置 + {n_dt} 行数据行模板 + {n_ft} 行底部行")
    return registry, data_templates


def merge_template_with_plm(template_rows: List[dict], plm_values: dict) -> List[dict]:
    """
    将 Excel 数据行模板与 PLM 含量值合并。
    
    template_rows: Excel 中的模板行列表（含显示属性，如 heavy, is_sub 等）
    plm_values:    PLM 提供的含量值字典，key=营养素英文名, value=含量值
                   如 {"Fat / Lipides": {"value": "1 g", "nrv": "1 %"}}
    
    返回: 合并后的 table_data（可直接传给渲染引擎）
    """
    result = []
    for tmpl in template_rows:
        row = dict(tmpl)  # 复制模板属性
        name = row.get("name", "")
        
        # 从 PLM 查找对应的含量值
        if name in plm_values:
            plm = plm_values[name]
            if isinstance(plm, dict):
                row.update(plm)  # 合并 PLM 的 value, nrv, per_serving 等
            else:
                row["value"] = str(plm)
        
        result.append(row)
    return result


# ══════════════════════════════════════════════════════════
# 入口
# ══════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python nut_config_excel.py export [output.xlsx]")
        print("  python nut_config_excel.py load [input.xlsx]")
        return
    
    cmd = sys.argv[1]
    
    if cmd == "export":
        path = sys.argv[2] if len(sys.argv) > 2 else "nut_config.xlsx"
        export_to_excel(path)
    elif cmd == "load":
        path = sys.argv[2] if len(sys.argv) > 2 else "nut_config.xlsx"
        registry, data_templates = load_from_excel(path)
        for code, layout in registry.items():
            dt_count = len(data_templates.get(code, []))
            print(f"  {code}: {layout.name} ({len(layout.header_rows)} header, {len(layout.columns)} cols, {dt_count} data rows)")
    else:
        print(f"未知命令: {cmd}")


if __name__ == "__main__":
    main()
